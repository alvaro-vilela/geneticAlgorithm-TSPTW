from openpyxl import load_workbook, Workbook
import numpy as np
from math import sqrt
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill, Border, Side


# Input from excel file
def Input(fname):
    # global wb
    wb = load_workbook(fname, data_only=True)
    # global ws
    ws = wb.worksheets[0]

    # global n,cox, coy, a, b, pen

    n = 0
    i = 2
    while ws.cell(i,1).value is not None:
       n+=1
       i+=1

    cox = np.zeros(n,dtype=float)
    coy = np.zeros(n,dtype=float)
    a = np.zeros(n,dtype=float)
    b = np.zeros(n,dtype=float)
    pen = np.zeros(n,dtype=float)

    for i in range(n):
        cox[i] = ws.cell(2+i, 2).value
        coy[i] = ws.cell(2+i, 3).value
        a[i] = ws.cell(2+i, 4).value
        b[i] = ws.cell(2+i, 5).value
        pen[i] = ws.cell(2+i, 6).value

    # global dist
    dist=np.zeros((n,n),dtype=float)

    # global dmax
    dmax=0

    for i in range(n):
        for j in range(n):
            dist[i][j]= sqrt((cox[i]-cox[j])**2+ (coy[i]-coy[j])**2)
            dmax+= dist[i][j]
    
    return wb, ws, n, pen, dist, dmax, a, b, cox, coy, a, b, dist, dmax

"""## Auxiliar functions"""

def Plot_graph(n, cox, coy):
  for i in range(1, n):
    plt.plot(cox[i],coy[i],'.k', ms = 5) # node
    plt.text(cox[i]+0.3, coy[i]+0.5, i) # node label
  plt.plot(cox[0], coy[0], 'sw', ms =6, mec = 'k') # warehouse
  plt.show()

def Plot_route(cox, coy, route):
  for i in range(len(route)-1):
    plt.plot([cox[route[i]], cox[route[i+1]]], [coy[route[i]], coy[route[i+1]]], 'k')
    plt.text(cox[route[i]]-0.3, coy[route[i]]+0.5, route[i])  if route[i] != 0 else None
  plt.plot(cox[0], coy[0], 'sw', ms = 10, mec = 'k') # warehouse
  plt.show()

def Calculate_distance(dist, route):
  distance=0

  for i in range(1,len(route)):
    distance += dist[route[i-1]][route[i]]
  return distance

def Calculate_cost(dist, a, b, pen, route):
    cost=0
    t_a = 0
    t_d = 0

    for i in range(1,len(route)):
      cost += dist[route[i-1]][route[i]]
      # print(f"Cost os moving {route[i-1]} -> {route[i]} = {dist[route[i-1]][route[i]]}")

      t_a = t_d + dist[route[i-1]][route[i]]

      if t_a < a[route[i]]:
        t_d = a[route[i]]
        # print(f"Waiting time = {round(t_d - t_a, 2)}")

      elif t_a > b[route[i]]:
        t_d = t_a
        # print(f"Retard = {round(t_a - b[route[i]], 2)}")
        # print(f"Cost of retard = {pen[route[i]] * (t_a - b[route[i]])}")
        cost += pen[route[i]] * (t_a - b[route[i]])
      else:
        t_d = t_a

      # print(f"Cumulative cost = {cost}")
      # print(f"t_d = {round(t_d, 2)}\n-----------------------------")

    return cost

def Download_excel(name, POP_SIZE, generations, n, dist, a, b, pen):
  wb = Workbook()
  ws = wb.active
  ws.title = "Initial Population"

  background_color = PatternFill(start_color="ADD8E6",  # Hex code for light blue
                                  end_color="ADD8E6",
                                  fill_type="solid")
  thin_border = Border(
      right=Side(border_style="thin", color="000000")  # "thin" style with black color
  )

  ws.column_dimensions['A'].width = 20
  ws.cell(1, 1).value = "ROute cost"
  ws.cell(1, 1).fill = background_color
  ws.cell(1, 1).border = thin_border
  ws.cell(1, 2).value = "Route"
  ws.cell(1, 2).fill = background_color

  for i in range(POP_SIZE):
    ws.cell(2+i, 1).value = Calculate_cost(dist, a, b, pen, generations[0][i])
    ws.cell(2+i, 1).border = thin_border
    for j in range(n+1):
      ws.cell(2+i, 2+j).value = generations[0][i][j]

  # ----------------------------------------------------

  new_sheet = wb.create_sheet(title="Final Population")
  new_sheet.column_dimensions['A'].width = 20
  new_sheet.cell(1, 1).value = "Route cost"
  new_sheet.cell(1, 1).fill = background_color
  new_sheet.cell(1, 1).border = thin_border
  new_sheet.cell(1, 2).value = "Route"
  new_sheet.cell(1, 2).fill = background_color

  for i in range(POP_SIZE):
    new_sheet.cell(2+i, 1).value = Calculate_cost(dist, a, b, pen, generations[0][i])
    new_sheet.cell(2+i, 1).border = thin_border
    for j in range(n+1):
      new_sheet.cell(2+i, 2+j).value = generations[-1][i][j]

  wb.save(name+".xlsx")